import requests
from lxml import html
import openpyxl
from tqdm import tqdm
import random
from fake_useragent import UserAgent

ua = UserAgent()
headers = {
    "User-Agent": ua.random
}
print(headers)
USER_AGENTS = [
    ua.random,
    ua.random,
    ua.random,
    ua.random,
    ua.random,
    ua.random,
    # 更多的User-Agent
]

def extract_content_from_url(url, max_retries=5):
    for _ in range(max_retries):
        headers = {
            "User-Agent": random.choice(USER_AGENTS)
        }
    response = requests.get(url, headers=headers)
    if response.status_code == 200:
        tree = html.fromstring(response.content)
        details_element = tree.xpath('//*[@id="productDetails_detailBullets_sections1"]')
        if not details_element:
            details_element = tree.xpath('//*[@id="detailBulletsWrapper_feature_div"]')
        if not details_element:
            return None
        details = details_element[0].xpath('.//text()')
        return ' '.join([item.strip() for item in details if item.strip()])
    else:
        print(f"Attempt {(_ + 1)}: Failed to get content from {url}. Status code: {response.status_code}")
    return None


def extract_asin(content):
    asin_index = content.find("ASIN")
    if asin_index != -1:
        return content[asin_index:].split()[1]
    return None


def extract_bestseller_ranks(content):
    rank_start_phrases = {
        "English": "Best Sellers Rank",
        "UK": "Best Sellers Rank",
        "German": "Amazon Bestseller-Rang",
        "French": "Classement des meilleures ventes d'Amazon",
        "Italian": "Posizione nella classifica Bestseller di Amazon",
        "ES": "Clasificación en los más vendidos de Amazon"
    }

    # 判断语言类型
    language = None
    rank_index = -1
    for lang, phrase in rank_start_phrases.items():
        rank_index = content.find(phrase)
        if rank_index != -1:
            language = lang
            break

    # 如果找到了其中一个前缀
    if rank_index != -1:
        ranks_section = content[rank_index:]
        if language == "English":
            # ranks_section = ranks_section.split("Date First Available", "Customer reviews")[0]

            # 找到两个分隔符的位置
            pos1 = ranks_section.find("Date First Available")
            pos2 = ranks_section.find("Customer reviews")

            # 如果找不到分隔符，将位置设为一个大值，这样不会影响到后面的min函数
            if pos1 == -1:
                pos1 = float('inf')
            if pos2 == -1:
                pos2 = float('inf')

            # 确定两个分隔符中较先出现的一个的位置
            first_delimiter_pos = min(pos1, pos2)

            # 根据找到的分隔符的位置来截取字符串
            if first_delimiter_pos != float('inf'):
                ranks_section = ranks_section[:first_delimiter_pos]

            ranks = [int(rank.split(" in")[0].strip().replace(',', '').replace('.', '')) for rank in
                     ranks_section.split() if rank.split(" in")[0].strip().replace(',', '').replace('.', '').isdigit()]
            if "See Top 100 in" in ranks_section:
                ranks = [rank for rank in ranks if rank != 100]

        elif language == "French":
            for split_key in ["Date de mise en", "Commentaires client"]:
                if split_key in ranks_section:
                    ranks_section = ranks_section.split(split_key)[0]
                    break
            ranks = [int(rank.split("en")[0].strip().replace(',', '').replace('.', '')) for rank in
                     ranks_section.split() if rank.split("en")[0].strip().replace(',', '').replace('.', '').isdigit()]
            if "Voir les 100 premiers en" in ranks_section:
                ranks = [rank for rank in ranks if rank != 100]

        elif language == "Italian":
            for split_key in ["Restrizioni di spedizione", "Recensioni dei clienti",
                              "Disponibile su Amazon.it a partire dal"]:
                if split_key in ranks_section:
                    ranks_section = ranks_section.split(split_key)[0]
                    break
            ranks = [int(rank.split("in")[0].strip().replace(',', '').replace('.', '')) for rank in
                     ranks_section.split() if rank.split("in")[0].strip().replace(',', '').replace('.', '').isdigit()]
            if "Visualizza i Top 100 nella categoria" in ranks_section:
                ranks = [rank for rank in ranks if rank != 100]

        elif language == "ES":
            for split_key in ["Restricciones de envío", "Opiniones de los clientes", "Producto en Amazon.es desde"]:
                if split_key in ranks_section:
                    ranks_section = ranks_section.split(split_key)[0]
                    break
            ranks = [int(''.join(filter(str.isdigit, rank.split("nº")[1]))) for rank in ranks_section.split("en") if
                     "nº" in rank]
            if "Top 100 en" in ranks_section:
                ranks = [rank for rank in ranks if rank != 100]

        else:  # This would be for German and other unspecified languages
            ranks = []
            for rank_str in ["Nr.", "Rank"]:
                ranks += [int(rank.split("in")[0].strip().replace(',', '').replace('.', '')) for rank in
                          ranks_section.split(rank_str)[1:] if
                          rank.split("in")[0].strip().replace(',', '').replace('.', '').isdigit()]

        return ranks
    return []


def update_excel_with_ranks(file_name, links):
    workbook = openpyxl.load_workbook(file_name)
    for sheet in workbook.worksheets:
        link_column_index = None
        for cell in sheet[2]:  # 查找链接列
            if cell.value == "链接":
                link_column_index = cell.column
                break
        if link_column_index:
            for row in tqdm(sheet.iter_rows(min_col=link_column_index, max_col=link_column_index, min_row=3),
                            desc="Processing links", total=sheet.max_row - 2):

                # for row in sheet.iter_rows(min_col=link_column_index, max_col=link_column_index, min_row=3):
                link_cell = row[0]
                if link_cell.value and link_cell.value in links:
                    content = extract_content_from_url(link_cell.value)
                    if content is None:
                        print(f"Unable to extract content from {link_cell.value}")
                        continue  # 如果内容为None，跳过当前链接处理
                    ranks = extract_bestseller_ranks(content)
                    # ranks_numeric = [''.join(filter(str.isdigit, rank)) for rank in ranks]
                    ranks_numeric = [str(rank) for rank in ranks]
                    for idx, rank in enumerate(ranks_numeric, start=1):
                        sheet.cell(row=link_cell.row, column=link_column_index + idx + 2).value = rank
    workbook.save(file_name)


def process_links(file_name, links):
    update_excel_with_ranks(file_name, links)


if __name__ == "__main__":
    # 替换下面的URL为您想要测试的URL
    test_url = 'https://www.amazon.co.uk/dp/B09YLSPS91?ref=myi_title_dp'

    content = extract_content_from_url(test_url)
    print(content)

    if content:
        asin_value = extract_asin(content)
        rank_values = extract_bestseller_ranks(content)

        if asin_value:
            print("ASIN:", asin_value)
        else:
            print("Unable to extract ASIN from content.")

        if rank_values:
            for idx, rank in enumerate(rank_values, 1):
                print(f"Rank {idx}:", rank)
        else:
            print("Unable to extract Rank from content.")
    else:
        print("Failed to retrieve or extract details from the provided URL.")
